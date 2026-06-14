import openpyxl
from openpyxl.chart import LineChart, Reference

wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=2, value='d1')
ws.cell(row=1, column=3, value='d2')

for i in range(1, 6):
    ws.cell(row=i+1, column=1, value=f"X{i}")
    ws.cell(row=i+1, column=2, value=i * i)
    ws.cell(row=i+1, column=3, value=i % 2)

chart = LineChart()
chart.title = 'Sample Chart'
chart.x_axis.title = 'X'
chart.y_axis.title = 'Y'

data = Reference(ws, min_row=1, min_col=2, max_row=6, max_col=3)
chart.add_data(data, titles_from_data=True)

labels = Reference(ws, min_col=1, min_row=2, max_row=6)
chart.set_categories(labels)

chart.legend = None

ws.add_chart(chart, 'E1')

wb.save('output_LineChart.xlsx')